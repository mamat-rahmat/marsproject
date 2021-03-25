from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font
from django.http import HttpResponse
from django.contrib import admin
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.flatpages.models import FlatPage
from django.contrib.flatpages.admin import FlatPageAdmin as FlatPageAdminOld
from django.contrib.flatpages.admin import FlatpageForm as FlatpageFormOld
from django.contrib.auth.models import User
from django import forms
from django_summernote.widgets import SummernoteWidget
from .models import Program, SelectedProgram, UserProfile, Tutorship, Membership, Problemset, Exam, Answer, Topic, Post


class FlatpageForm(FlatpageFormOld):
    content = forms.CharField(widget=SummernoteWidget())

    class Meta:
        model = FlatPage
        fields = '__all__'

class FlatPageAdmin(FlatPageAdminOld):
    form = FlatpageForm


# We have to unregister the normal admin, and then reregister ours
admin.site.unregister(FlatPage)
admin.site.register(FlatPage, FlatPageAdmin)


# Define a new User admin
class UserAdmin(BaseUserAdmin):
    search_fields = ['username']

# Re-register UserAdmin
admin.site.unregister(User)
admin.site.register(User, UserAdmin)


class ExamInline(admin.TabularInline):
    model = Exam
    extra = 0

class ProgramAdmin(admin.ModelAdmin):
    inlines = [ExamInline]
    list_display = ('name', 'bidang', 'exam_count')
    search_fields = ['name']
    actions = ['clone_program', 'export_ranking']

    def clone_program(self, request, queryset):
        for obj in queryset:
            # Get exams
            exams = obj.exam_set.all()
            # Clone program
            obj.pk = None
            obj.name += " (copy)"
            obj.save()
            # Clone exams and refer to cloned program
            for exam in exams:
                exam.pk = None
                exam.program = obj
                exam.save()
        self.message_user(request, "%s programs succesfully cloned" % queryset.count())

    def export_ranking(self, request, queryset):
        wb = Workbook()

        for program in queryset:
            exams = program.exam_set.all()
            members_siswa = Membership.objects.select_related('user__userprofile').filter(program=program, user__userprofile__role='SISWA')
            members_guru = Membership.objects.select_related('user__userprofile').filter(program=program, user__userprofile__role='GURU')
            header = ['Role', 'Nama', 'Sekolah', 'Bidang', *[exam.name for exam in exams], 'Total Nilai', 'Total Waktu']
            body = []
            for member in members_siswa:
                user = member.user
                userprofile = user.userprofile
                row = ['SISWA', userprofile.nama_lengkap, userprofile.sekolah, userprofile.get_bidang_display()]
                total = 0
                waktu = timedelta()
                for exam in exams:
                    answer, created = Answer.objects.get_or_create(user=user, exam=exam)
                    row.append(answer.score)
                    total += answer.score
                    waktu += answer.updated_at - exam.start_time
                row.append(total)
                row.append(waktu)
                body.append(row)
            for member in members_guru:
                userprofile = member.user.userprofile
                row = ['GURU', userprofile.nama_lengkap, userprofile.sekolah, userprofile.get_bidang_display()]
                total = 0
                waktu = timedelta()
                for exam in exams:
                    answer, created = Answer.objects.get_or_create(user=user, exam=exam)
                    row.append(answer.score)
                    total += answer.score
                    waktu += answer.updated_at - exam.start_time
                row.append(total)
                row.append(waktu)
                body.append(row)
            body.sort(key=lambda x: (x[0],x[-2],-x[-1]), reverse=True)

            ws = wb.create_sheet()
            ws.append([program.name])
            ws.append(header)
            for row in body:
                ws.append(row)
            c = ws['A1']
            c.font = Font(bold=True)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=ranking.xlsx'

        sheet = wb.get_sheet_by_name('Sheet')
        wb.remove_sheet(sheet)
        wb.save(response)
        return response

admin.site.register(Program, ProgramAdmin)



class SelectedProgramAdmin(admin.ModelAdmin):
    list_display = ('program',)

admin.site.register(SelectedProgram, SelectedProgramAdmin)



class UserProfileAdmin(admin.ModelAdmin):
    list_display = ('id', 'user', 'nama_lengkap', 'role', 'bidang', 'sekolah')
    search_fields = ['nama_lengkap', 'sekolah']
    list_filter = ('role',)
    actions = ['add_to_selected_program']

    def add_to_selected_program(self, request, queryset):
        for obj in queryset:
            for selected in SelectedProgram.objects.all():
                membership, created = Membership.objects.get_or_create(
                    user=obj.user,
                    program=selected.program
                )
                membership.paid = True
                membership.save()
        self.message_user(request, "%s userprofiles succesfully added to selected program" % queryset.count())


admin.site.register(UserProfile, UserProfileAdmin)


class TutorshipAdmin(admin.ModelAdmin):
    list_display = ('user', 'program')

admin.site.register(Tutorship, TutorshipAdmin)


class MembershipAdmin(admin.ModelAdmin):
    list_display = ('user', 'program', 'paid')
    #list_display = ('user', 'user__userprofile__nama_lengkap', 'user__userprofile_sekolah', 'program', 'paid')
    search_fields = ['user__username', 'program__name']
    actions = ['accept_membership']
    list_filter = ['paid', 'program__bidang']

    def accept_membership(self, request, queryset):
        for obj in queryset:
            obj.paid = True
            obj.save()
        self.message_user(request, "%s membership succesfully accepted" % queryset.count())

admin.site.register(Membership, MembershipAdmin)


class ProblemsetAdmin(admin.ModelAdmin):
    list_display = ('name','mcq_total')
    fieldsets = [
        (None, {'fields': ['name']}),
        ('Problem & Solution', {'fields': ['problem_file', 'solution_file']}),
        ('Answer Key', {'fields': [ ('no01', 'no11', 'no21', 'no31', 'no41'),
                                    ('no02', 'no12', 'no22', 'no32', 'no42'),
                                    ('no03', 'no13', 'no23', 'no33', 'no43'),
                                    ('no04', 'no14', 'no24', 'no34', 'no44'),
                                    ('no05', 'no15', 'no25', 'no35', 'no45'),
                                    ('no06', 'no16', 'no26', 'no36', 'no46'),
                                    ('no07', 'no17', 'no27', 'no37', 'no47'),
                                    ('no08', 'no18', 'no28', 'no38', 'no48'),
                                    ('no09', 'no19', 'no29', 'no39', 'no49'),
                                    ('no10', 'no20', 'no30', 'no40', 'no50')]}),
        ('Essay Question', {'fields': ['add_essay']}),
    ]
    class Media:
        css = {
            'all': ('core/css/label-admin.css',)
        }

admin.site.register(Problemset, ProblemsetAdmin)


class AnswerAdmin(admin.ModelAdmin):
    list_display = ('user', 'exam', 'graded', 'score')
    fieldsets = [
        (None, {'fields': ['user', 'exam', 'graded', 'score']}),
        ('Answer Choice', {'fields': [  ('no01', 'no11', 'no21', 'no31', 'no41'),
                                        ('no02', 'no12', 'no22', 'no32', 'no42'),
                                        ('no03', 'no13', 'no23', 'no33', 'no43'),
                                        ('no04', 'no14', 'no24', 'no34', 'no44'),
                                        ('no05', 'no15', 'no25', 'no35', 'no45'),
                                        ('no06', 'no16', 'no26', 'no36', 'no46'),
                                        ('no07', 'no17', 'no27', 'no37', 'no47'),
                                        ('no08', 'no18', 'no28', 'no38', 'no48'),
                                        ('no09', 'no19', 'no29', 'no39', 'no49'),
                                        ('no10', 'no20', 'no30', 'no40', 'no50'), 
                                     ]})
    ]
    actions = ['grade_answers']
    search_fields = ['user__username', 'exam__name', 'exam__program__name']

    class Media:
        css = {
            'all': ('core/css/label-admin.css',)
        }

    def grade_answers(self, request, queryset):
        for obj in queryset:
            obj.grade()
        self.message_user(request, "%s answers succesfully graded" % queryset.count())

admin.site.register(Answer, AnswerAdmin)


class TopicAdmin(admin.ModelAdmin):
    list_display = ('title', 'author')

admin.site.register(Topic, TopicAdmin)


class PostAdmin(admin.ModelAdmin):
    list_display = ('id', 'topic', 'author')

admin.site.register(Post, PostAdmin)

