from openpyxl import load_workbook
from core.models import UserProfile, Program, Membership
from django.contrib.auth.models import User


getbidang = {
    'MATEMATIKA': 'matematika-smp',
    'IPA': 'ipa-smp',
    'IPS': 'ips-smp',
}

getprogram = {
    'matematika-smp': Program.objects.get(bidang='matematika-smp', name__startswith='MARS'),
    'ipa-smp': Program.objects.get(bidang='ipa-smp', name__startswith='MARS'),
    'ips-smp': Program.objects.get(bidang='ips-smp', name__startswith='MARS'),
    'Testing': Program.objects.get(name__startswith='Testing')
}


filename = '26mar.xlsx'
wb = load_workbook(filename)
ws = wb.active
for row in list(ws)[1:]:
    rownum = row[0].row
    nama = row[1].value
    sekolah = row[2].value
    bidang = getbidang[row[3].value]
    password = User.objects.make_random_password(length=8)
    user = User.objects.create_user(
        username = 'temp',
        password = password,
        email = 'temp@example.com'
    )
    user.username = "mars{:03d}".format(user.id)
    user.email = "mars{:03d}@example.com".format(user.id)
    user.save()
    ws['E'+str(rownum)] = user.username
    ws['F'+str(rownum)] = password
    userprofile = UserProfile.objects.create(
        nama_lengkap = nama,
        user = user,
        nomor_whatsapp = '-',
        sekolah = sekolah,
        kota = '-',
        provinsi = '-',
        bidang = bidang,
        guru_pembimbing = '-',
        nomor_guru_pembimbing = '-',
    )
    membership = Membership.objects.create(
        user = user,
        program = getprogram[bidang],
        paid = True
    )
    membership = Membership.objects.create(
        user = user,
        program = getprogram['Testing'],
        paid = True
    )

wb.save(filename = filename)
